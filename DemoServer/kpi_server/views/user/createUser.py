# coding=utf8
from rest_framework.decorators import api_view

from django.http.response import JsonResponse

from kpi_server.models import Users

import datetime
from openpyxl import load_workbook

'''新增用户'''
@api_view(['POST'])
def createUser(request):

    # body解析
    body_data = {
        'is_multi':request.data.get('is_multi',False),
        'update_data':request.data.get('update_data',None),
        'update_file':request.FILES.get('update_file',None)
    }

    print(body_data)

    # 查询提交的内容
    if body_data['is_multi'] and body_data['update_file'] is not None:

        # excel文件校验
        # 解析文件
        xls_files = load_workbook(body_data['update_file'])
        xls_sheet = xls_files['用户导入模板']

        user_list = []
        for row in xls_sheet.iter_rows(min_row=7,values_only=True):
            each_user = {
                'notes_id' : row[0],
                'user_name': row[1],
                'user_character':int(row[2].split('-')[0]),
                'user_belong_group':int(row[3].split('-')[0]),
                'user_belong_org':int(row[4]),
                'user_state':0,
                'user_create': datetime.date.today().strftime("%Y-%m-%d"),
                'user_update': datetime.date.today().strftime("%Y-%m-%d"),
            }
            user_list.append(each_user)

        # try:
        #     # 批量添加
        #     Users.objects.bulk_create(user_list)

        # except IntegrityError as e:
        #     # 唯一约束错误
        #     print('唯一约束错误',e)

        err_row_list = []

        for index,each_user in enumerate(user_list):
            obj,created = Users.objects.get_or_create(**each_user)
            # 插入失败的存入err_row_list
            if created:
                pass
            else:
                err_row_list.append(index)

        re_msg = {'code':1,'err':err_row_list,'msg':'done'}

    else:
        # 查询notesid是否存在，

        try:
            user_obj = Users.objects.get(notes_id=body_data.get('update_data').get('notes_id'))
            re_msg = {'code':'2','msg':'users exists.'}

        except Users.MultipleObjectsReturned:
            re_msg = {'code':3,'msg':'multi users exists'}

        except Users.DoesNotExist:
            new_user = Users()
            for (x,y) in body_data.get('update_data').items():
                setattr(new_user,x,y)
            new_user.save()
            re_msg = {'code':0,'msg':'create_success'}

    return JsonResponse(re_msg,safe=False)