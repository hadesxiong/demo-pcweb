# coding=utf8
from rest_framework.decorators import api_view,permission_classes
from rest_framework.permissions import IsAuthenticated

from django.http.response import JsonResponse

from kpi_server.models import IndexDetail,Index,UploadRecord,UploadDetail,Org
from kpi_server.serializers import UploadDetailSerializer

import pandas as pd
import datetime,hashlib
from openpyxl import load_workbook
from io import BytesIO

# 通用方法 - 获取md5
def get_md5(file):
    md5_hash = hashlib.md5()
    for chunk in file.chunks():
        md5_hash.update(chunk)
    return md5_hash.hexdigest()

# 通用方法 - 解析uploadDetail
def readDetail(file,detail_id,index_dict,org_dict):

    # 解析文件
    xls_files = load_workbook(file)
    xls_sheet = xls_files['导入数据模板']

    # 形成列表
    index_list = []
    for row in xls_sheet.iter_rows(min_row=0,values_only=True):
        index_list.append(row)

    # 处理df
    df = pd.DataFrame(index_list[1:],columns=index_list[0]).reset_index(drop=True)

    # 更换指标名称
    df = df.rename(columns=index_dict,errors='ignore')
    df = df.rename(columns={'机构名称':'detail_belong','指标日期':'detail_date','是否计划':'detail_type'})

    # melt后更换机构名称
    df = df.melt(
        id_vars=['detail_belong', 'detail_date', 'detail_type'],
        var_name='index_num',
        value_name='detail_value'         
    )
    df['detail_belong'] = df['detail_belong'].replace(org_dict)
    df['detail_id'] = detail_id
    
    return df.to_dict(orient='records')


# 接口方法 - 上传数据
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createUpload(request):

    # 解析body_data
    body_data = {
        'record_class': int(request.data.get('class',None)),
        'record_date': request.data.get('date',None),
        'update_file': request.FILES.get('file',None),
        'update_user': request.data.get('user',None)
    }

    if None in body_data.values():
        re_msg = {'code':1,'msg':'err msg.'}
    
    else:
        # 标准部分处理
        now_date = datetime.datetime.now()
        detail_id = f"RD_{now_date.strftime('%Y%m%d%H%M')}"

        # 处理指标字典
        index_queryset = Index.objects.all().values('index_name','index_num')
        index_dict = pd.DataFrame(index_queryset).set_index('index_name')['index_num'].to_dict()

        # 处理机构字典
        org_queryset = Org.objects.all().values('org_name','org_num')
        org_dict = pd.DataFrame(org_queryset).set_index('org_name')['org_num'].to_dict()

        # 判断class与date是否存在对应值，如果存在，则为更新，反之则为创建
        try:
            record_queryset = UploadRecord.objects.get(
                record_class= body_data['record_class'],
                record_date = body_data['record_date']
            )

            old_active = UploadDetail.objects.get(record_id=record_queryset.record_id,detail_active=1)

            newactive_obj = {
                'detail_id':detail_id,
                'record_id': record_queryset.record_id,
                'detail_update_user': body_data['update_user'],
                'detail_state':1,
                'detail_create': now_date.strftime("%Y-%m-%d"),
                'detail_update': now_date.strftime("%Y-%m-%d"),
                'detail_update_fileName':body_data['update_file'].name,
                'detail_update_fileMD5':get_md5(body_data['update_file']),
                'detail_ext_info':'',
                'detail_active':1
            }

            # 解析文件

            # new_active = UploadDetail(**newactive_obj)
            # new_active.save()

            # old_active.detail_active = 0
            # old_active.save()

            # 获取old_active的detail_id,去Index_detail匹配,并更新掉全部的detail_state
            oldIndex_queryset = IndexDetail.objects.filter(detail_id=old_active.detail_id,detail_state=1).values('id','detail_state')
            oldIndex_queryset.update(detail_state=0)

            # 把解析的数据存入IndexDetail
            newIndex_data = readDetail(file=body_data['update_file'],detail_id=detail_id,index_dict=index_dict,org_dict=org_dict)
            IndexDetail.objects.bulk_create(newIndex_data)

            re_msg = {'code':0,'msg':'update'}

        except UploadRecord.DoesNotExist:

            # 新建数据并保存
            # 新建record
            record_id = f"UR_{now_date.strftime('%Y%m%d%H%M')}"
            newRecord_obj = {
                'record_id':record_id,
                'record_class':body_data['record_class'],
                'record_date': now_date.strftime("%Y-%m-%d"),
                'record_update_user': body_data['update_user'],
                'record_update_time': now_date.strftime("%Y-%m-%d %H-%M-%S"),
                'record_udpate_state':0,
                'record_crate': now_date.strftime("%Y-%m-%d %H-%M-%S"),
                'record_update': now_date.strftime("%Y-%m-%d %H-%M-%S"),
                'record_update_ext_info':''
            }

            new_record = UploadRecord(**newRecord_obj)
            new_record.save()

            newIndex_data = readDetail(file=body_data['update_file'],detail_id=detail_id,index_dict=index_dict,org_dict=org_dict)
            IndexDetail.objects.bulk_create(newIndex_data)

            re_msg = {'code':0,'msg':'create'}

    return JsonResponse(re_msg,safe=False)


# 接口方法 - 发布
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def publishUpload(request):

    # 解析body_data
    body_data = {
        'record_id': request.data.get('record',None),
        'detail_id': request.data.get('detail',None),
        'user': request.data.get('user',None)
    }

    if None in body_data.values():
        re_msg = {'code':1,'msg':'err msg'}

    else:
        # 获取当前在生效的detail_id
        # 将当前在生效的detail_id中的IndexDetail记录设为不生效
        # 将传入的detail_id设为生效
        oldDetail_detail_id = UploadDetail.objects.get(record_id=body_data['record_id'],detail_active=1).detail_id
        oldDetail_queryset = IndexDetail.objects.filter(detail_id=oldDetail_detail_id).values('detail_id','detail_state')
        oldDetail_queryset.update(detail_state=0)

        newDetail_queryset = IndexDetail.objects.filter(detail_id=body_data['detail_id']).values('detail_id','detail_state')
        newDetail_queryset.update(detail_state=1)

        # 设置record记录为生效
        record_obj = UploadRecord.objects.filter(record_id=body_data['record_id']).values('')
        update_datetime = datetime.datetime.now().strftime("%Y-%m-%d")
        record_obj.update(record_update_state=1,record_update=update_datetime)

        re_msg = {'code':0,'msg':'success'}

    return JsonResponse(re_msg,safe=False)