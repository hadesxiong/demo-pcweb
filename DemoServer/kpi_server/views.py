from django.shortcuts import render

from django.core.paginator import Paginator
from django.db.models import Subquery,Q
from django.db import IntegrityError,connections

from django.http.response import JsonResponse
from rest_framework.decorators import api_view

from kpi_server.models import Users,Org,Reference,Index,IndexDetail

from kpi_server.serializers import UsersSerializer,OrgSerializer,RefSerializer,IndexSerializer,IndexDetailRankSerializer

import math,datetime,itertools

from openpyxl import load_workbook
from decimal import Decimal
# Create your views here.

'''查询用户列表'''

@api_view(['GET'])
def getUserList(request):

    # params解析
    query_params = {
        'belong_group': int(request.query_params.get('group',0)),
        'user_character': int(request.query_params.get('character',0)),
        'org_level': int(request.query_params.get('org',0)),
        'key_word':request.query_params.get('ext',None),
        'user_client': request.query_params.get('client',None),
        'page_size':int(request.query_params.get('size',15)),
        'page':int(request.query_params.get('page',1))
    }
    
    # 筛选出机构层级
    if query_params['org_level'] == 0:
        org_query = Org.objects.all().values('org_num','org_name')
    else:
        org_query = Org.objects.filter(org_level=query_params['org_level']).values('org_num','org_name')

    # 关键词
    if query_params['key_word'] is not None:
        complex_condition = Q(notes_id__contains=query_params['key_word']) | Q(user_name__contains=query_params['key_word']) | Q(user_belong_org__in=Subquery(Org.objects.filter(org_name__contains=query_params['key_word']).values('org_num')))
    else:
        complex_condition = Q()

    # 筛选用户所属条线
    if query_params['belong_group'] == 0:
        group_condition = Q()        
    else:
        group_condition = Q(user_belong_group=query_params['belong_group'])

    # 筛选用户角色
    if query_params['user_character'] == 0:
        character_condition = Q()
    else:
        character_condition = Q(user_character=query_params['user_character'])

    users_querySet = Users.objects.filter(user_belong_org__in=Subquery(org_query.values('org_num'))).filter(group_condition).filter(character_condition).filter(complex_condition)

    # 分页
    page_inator = Paginator(users_querySet,query_params['page_size'])
    page_max = math.ceil(len(users_querySet)/query_params['page_size'])

    if query_params['page'] <= page_max:
        each_users_list = page_inator.page(query_params['page'])
        users_serializer = UsersSerializer(each_users_list,many=True)
        re_msg = {'data':users_serializer.data,'code':0,'msg':'success','has_next':(query_params['page']<page_max)}
    else:
        re_msg = {'code':1,'msg':'error_range'}

    return JsonResponse(re_msg,safe=False)

'''删除/更新用户信息'''

@api_view(['POST'])
def updateUser(request):

    # body解析
    body_data = {
        'type':request.data.get('type','update'),
        'notes_id':request.data.get('user',None),
        'update_data':request.data.get('update_data',None),
    }

    if body_data['notes_id'] is not None:
        user_obj = Users.objects.get(notes_id=body_data['notes_id'])

        if user_obj and  body_data['update_data'] is not None and (body_data['type'] == 'update'):
            for (x,y) in body_data['update_data'].items():
                print(x,y)
                setattr(user_obj,x,y)
            user_obj.save()
            re_msg = {'code':0,'msg':'update success'}

        elif user_obj and (body_data['type'] == 'delete'):
            user_obj.delete()
            re_msg = {'code':0,'msg':'delete success'}

        else:
            re_msg = {'code':1,'msg':'user not found'}

    else:
        re_msg = {'code':3,'msg':'error params.'}

    return JsonResponse(re_msg,safe=False)

'''新增用户'''
@api_view(['POST'])
def createUser(request):

    # body解析
    body_data = {
        'is_multi':request.data.get('is_multi',False),
        'update_data':request.data.get('update_data',None),
        'update_file':request.FILES.get('update_file',None)
    }

    print(body_data)

    # 查询提交的内容
    if body_data['is_multi'] and body_data['update_file'] is not None:

        # excel文件校验
        # 解析文件
        xls_files = load_workbook(body_data['update_file'])
        xls_sheet = xls_files['用户导入模板']

        user_list = []
        for row in xls_sheet.iter_rows(min_row=7,values_only=True):
            each_user = {
                'notes_id' : row[0],
                'user_name': row[1],
                'user_character':int(row[2].split('-')[0]),
                'user_belong_group':int(row[3].split('-')[0]),
                'user_belong_org':int(row[4]),
                'user_state':0,
                'user_create': datetime.date.today().strftime("%Y-%m-%d"),
                'user_update': datetime.date.today().strftime("%Y-%m-%d"),
            }
            user_list.append(each_user)

        # try:
        #     # 批量添加
        #     Users.objects.bulk_create(user_list)

        # except IntegrityError as e:
        #     # 唯一约束错误
        #     print('唯一约束错误',e)

        err_row_list = []

        for index,each_user in enumerate(user_list):
            obj,created = Users.objects.get_or_create(**each_user)
            # 插入失败的存入err_row_list
            if created:
                pass
            else:
                err_row_list.append(index)

        re_msg = {'code':1,'err':err_row_list,'msg':'done'}

    else:
        # 查询notesid是否存在，

        try:
            user_obj = Users.objects.get(notes_id=body_data.get('update_data').get('notes_id'))
            re_msg = {'code':'2','msg':'users exists.'}

        except Users.MultipleObjectsReturned:
            re_msg = {'code':3,'msg':'multi users exists'}

        except Users.DoesNotExist:
            new_user = Users()
            for (x,y) in body_data.get('update_data').items():
                setattr(new_user,x,y)
            new_user.save()
            re_msg = {'code':0,'msg':'create_success'}

    return JsonResponse(re_msg,safe=False)

'''查询机构列表'''
@api_view(['GET'])
def getOrgList(request):

    # params解析
    query_params = {
        'org_level': int(request.query_params.get('level',0)),
        'org_group': int(request.query_params.get('group',0)),
        'key_word': request.query_params.get('ext',None),
        'user_client': request.query_params.get('client',None),
        'page_size':int(request.query_params.get('size',15)),
        'page':int(request.query_params.get('page',1))
    }

    # 分组筛选条件
    if query_params['org_group'] == 0:
        group_condition = Q()
    else:
        group_condition = Q(org_group=query_params['org_group'])

    # 层级筛选条件
    if query_params['org_level'] == 0:
        level_condition = Q()
    else:
        level_condition = Q(org_level=query_params['org_level'])

    # 关键词筛选
    if query_params['key_word'] is not None:
        complex_condition = Q(org_name__contains=query_params['key_word']) | Q(org_num__contains=query_params['key_word']) | Q(org_manager__in=Subquery(Users.objects.filter(notes_id__contains=query_params['key_word']).values('notes_id'))) | Q(org_manager__in=Subquery(Users.objects.filter(notes_id__contains=query_params['key_word']).values('notes_id')))
    else:
        complex_condition = Q()

    org_querySet = Org.objects.filter(group_condition).filter(level_condition).filter(complex_condition)

    # 分页
    page_inator = Paginator(org_querySet,query_params['page_size'])
    page_max = math.ceil(len(org_querySet)/query_params['page_size'])

    if query_params['page'] <= page_max:
        each_orgs_list = page_inator.page(query_params['page'])
        orgs_serializer = OrgSerializer(each_orgs_list,many=True)
        re_msg = {'data':orgs_serializer.data,'code':0,'msg':'success','has_next':(query_params['page']<page_max)}
    else:
        re_msg = {'code':1,'msg':'error_range'}

    return JsonResponse(re_msg,safe=False)

'''新增/更新/删除机构'''
@api_view(['POST'])
def updateOrg(request):

    body_data = {
        'update_type': request.data.get('type','create'),
        'org_num': request.data.get('num',None),
        'update_data': request.data.get('update_data',None)
    }

    # 新增逻辑
    if body_data['update_type'] == 'create':
        # 表单
        update_data = {
            'org_num':body_data['org_num'],
            'org_name':body_data['update_data']['org_name'],
            'parent_org_id':body_data['update_data']['parent_org'],
            'org_level':body_data['update_data']['org_level'],
            'org_group':body_data['update_data']['org_group'],
            'org_manager':body_data['update_data']['org_manager'],
            'org_state':0,
            'org_create':datetime.date.today().strftime("%Y-%m-%d"),
            'org_update':datetime.date.today().strftime("%Y-%m-%d")
        }
        # 检验是否为空
        if None in update_data.values():
            re_msg = {'code':1,'msg':'err params'}

        else:
            obj,created = Org.objects.get_or_create(**update_data)
            if created:
                re_msg = {'code':0,'msg':'created','err':''}
            else:
                re_msg = {'code':1,'msg':'exists','err':''}

    # 更新逻辑
    elif body_data['update_type'] == 'update' and body_data['update_data'] is not None:

        try:
            org_obj = Org.objects.get(org_num=body_data['org_num'])
            for (x,y) in body_data['update_data'].items():
                setattr(org_obj,x,y)
            
            org_obj.save()
            re_msg = {'code':0,'msg':'updated','err':''}
        
        except Org.DoesNotExist:
            re_msg = {'code':1,'msg':'org not exists'}

    elif body_data['update_type'] == 'delete':

        try:
            org_obj = Org.objects.get(org_num=body_data['org_num'])
            org_obj.delete()

            re_msg = {'code':0,'msg':'deleted','err':''}
        
        except Org.DoesNotExist:
            re_msg = {'code':1,'msg':'org not exists'}

    return JsonResponse(re_msg,safe=False)

'''筛选项处理-处筛选'''
@api_view(['GET'])
def getFilter(request):

    # params解析
    query_params = {
        'ref_type': request.query_params.get('type',None),
        'ref_scene': int(request.query_params.get('scene',0)),
        'client': request.query_params.get('client','client')
    }

    if None in query_params.values():
        re_msg = {'code':1,'msg':'error params'}
    else:
        # 拆分ref_type
        type_list = query_params['ref_type'].split('.')
        ref_queryset = Reference.objects.filter(ref_ext__in=type_list)

        ref_serializer = RefSerializer(ref_queryset,many=True)
        ref_result = {}
        for each_ref in ref_serializer.data:
            ref_type = each_ref['ref_type']
            if ref_type not in ref_result:
                # setattr(ref_result,ref_type,{'ref_code':each_ref['ref_code'],'ref_name':each_ref['ref_name']})
                ref_result[ref_type] = []
            ref_result[ref_type].append({'ref_code':each_ref['ref_code'],'ref_name':each_ref['ref_name']})

        print(ref_result)

        re_msg = {'code':0,'data':ref_result}

    return JsonResponse(re_msg,safe=False)

'''测试生成假数据'''
@api_view(['POST'])
def generateDetail(request):

    # 解析body
    body_data = {
        'detail_type':request.data.get('type',None),
        'detail_class':request.data.get('class',1),
        'detail_date':request.data.get('date',None),
        'detail_belong':request.data.get('belong','all'),
        'index_num':request.data.get('index','all')
    }

    if None in body_data.values():
        re_msg = {'code':1,'msg':'err params.'}

    else:

        # 指标
        if body_data['index_num'] == 'all':
            index_list = IndexSerializer(Index.objects.all(),many=True).data
            grouped_index = itertools.groupby(index_list,lambda x:x['index_num'])
        
        body_data['index_num'] = [item['index_num'] for _,group in grouped_index for item in group]

        # 机构
        if body_data['detail_belong'] == 'all':
            org_queryset = Org.objects.filter(Q(org_level__gte=3) & Q(org_level__lt=5))
            org_data = OrgSerializer(org_queryset,many=True).data
            body_data['detail_belong'] = list(map(lambda x:x['org_num'],org_data))

        # print(list(map(lambda x:x['org_num'],org_data)))


        body_data['detail_create'] = body_data['detail_date']
        body_data['detail_state'] = [0]
        body_data['detail_value'] = [0]

        keys = body_data.keys()
        values = list(itertools.product(*body_data.values()))

        result = [{k:v for k,v in zip(keys,value)} for value in values]

        # for each_result in result:
        #     IndexDetail.objects.get_or_create(**each_result)

        objects = [IndexDetail(**item) for item in result]
        IndexDetail.objects.bulk_create(objects)

        re_msg = {'code':0,'msg':'done'}

    return JsonResponse(re_msg,safe=False)

'''获取数据'''
# 获取数据接口较为负责，根据对应页面及功能进行区分
@api_view(['GET'])
def getRank(request):

    # params解析
    query_params = {
        # 'belong_line': request.query_params.get('line',0),
        'index_class': int(request.query_params.get('class',0)),
        'org_group': int(request.query_params.get('group',0)),
        'data_date': request.query_params.get('date',None)
    }

    # 处理日期
    if None in query_params.values():
        syym = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
        query_params['data_date'] = syym.strftime("%Y-%m-%d")

    # 指标筛选
    # 分类数据
    if query_params['index_class'] == 0:
        # index_condition = Q()
        # 后续补充动态
        index_condition = f'SELECT index_num FROM index_info WHERE index_class in (1,2,3,4,5)'
    else:
        # index_condition = Q(index_class=query_params['index_class'])
        index_condition = f'SELECT index_num FROM index_info WHERE index_class = {query_params["index_class"]}'

    # print(index_condition)

    # 是否重要数据补充
    
    # 机构筛选
    if query_params['org_group'] == 0:
        # group_condition = Q(org_group__gte=1) & Q(org_group__lte=3)
        group_condition = f'SELECT org_num FROM org_info WHERE org_group in (1,2,3)'
    else:
        # group_condition = Q(org_group=query_params['org_group'])
        group_condition = f'SELECT org_num FROM org_info WHERE org_group = {query_params["org_group"]}'

    # print(group_condition)

    # 获取计划筛选日期
    this_year = int(query_params['data_date'].split('-')[0])
    first_day = datetime.date(this_year,1,1).strftime("%Y-%m-%d")

    # 查询逻辑-sql
    set_var_query = '''
    SET @row_number := 0, @prev_index_num := NULL;
    '''
    
    get_data_query = f'''
    SELECT t.index_num, t.detail_belong, t.detail_date, t.detail_value AS result,
        (SELECT detail_value
            FROM index_detail AS d
            WHERE d.index_num = t.index_num
            AND d.detail_belong = t.detail_belong
            AND d.detail_date = '{first_day}'
            AND d.detail_type = 1
            ORDER BY d.detail_value DESC
            LIMIT 1) AS matching_detail_value
    FROM (
        SELECT 
            index_num, 
            detail_belong, 
            detail_date, 
            detail_value,
            CASE
                WHEN @prev_index_num = index_num THEN @row_number := @row_number + 1
                ELSE @row_number := 1
            END AS row_number,
            @prev_index_num := index_num
        FROM index_detail
        CROSS JOIN (SELECT @row_number := 0, @prev_index_num := NULL) AS vars
        WHERE (detail_date = '{query_params['data_date']}' 
            AND index_num IN ({index_condition}) 
            AND detail_belong IN ({group_condition}))
        ORDER BY index_num, detail_value DESC
    ) AS t
    WHERE t.row_number <= 5
    ORDER BY t.index_num, t.detail_value DESC;
    '''

    # 执行变量设定语句
    with connections['kpi_dashboard_db'].cursor() as cursor:
        cursor.execute(set_var_query)

    # 执行查询语句
    with connections['kpi_dashboard_db'].cursor() as cursor:
        cursor.execute(get_data_query)
        result = cursor.fetchall()

    # 结构化查询结果
    fetch_data = []
    for row in result:
        index_num,detail_belong,detail_date,detail_value,detail_plan = row
        fetch_data.append({
            'index_num':index_num,
            'detail_belong':detail_belong,
            'detail_date':detail_date,
            'detail_value':detail_value,
            'detail_plan':detail_plan,
            'detail_rate':round(detail_value/detail_plan,2)
        })
    # print(fetch_data)

    rank_data = IndexDetailRankSerializer(fetch_data,many=True).data



    return JsonResponse({'code':0,'data':rank_data},safe=False)